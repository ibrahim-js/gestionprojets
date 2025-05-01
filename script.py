import psycopg2
from openpyxl import load_workbook

# Database connection config
DB_CONFIG = {
    'dbname': 'gestionprojets',
    'user': 'postgres',
    'password': 'password',
    'host': '192.168.193.178',  # Use your server IP if needed
    'port': 5432
}
workbook = load_workbook(filename='cleanData.xlsx')
sheet = workbook.active

conn = psycopg2.connect(**DB_CONFIG)
cur = conn.cursor()

columns = [cell.value for cell in sheet[1]]
# Wrap column names with double quotes
quoted_columns = [f'"{col}"' for col in columns]

for row in sheet.iter_rows(min_row=2, values_only=True):
    placeholders = ', '.join(['%s'] * len(row))
    query = f"INSERT INTO projets2 ({', '.join(quoted_columns)}) VALUES ({placeholders})"
    cur.execute(query, row)

conn.commit()
cur.close()
conn.close()

print("✅ Data inserted successfully.")